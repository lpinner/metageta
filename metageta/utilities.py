# -*- coding: utf-8 -*-

# Copyright (c) 2013 Australian Government, Department of the Environment
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.

'''
Utility helper functions
'''
#========================================================================================================
# Imports
#========================================================================================================
import openpyxl
import sys, os.path, os, re, struct, glob, shutil,traceback,time,tempfile,copy
import warnings
import tarfile,zipfile
import uuid as _uuid

#========================================================================================================
# Globals
#========================================================================================================
dateformat='%Y-%m-%d'  #ISO 8601
timeformat='%H:%M:%S' #ISO 8601
datetimeformat='%sT%s' % (dateformat,timeformat)

encoding='utf-8'

iswin=os.name=='nt'#sys.platform[0:3].lower()=='win'#Are we on Windows

compressedfiles=('.zip','.tar.gz','.tgz','.tbz', '.tbz2','.tb2','.tar.bz2','.tar','kmz')


#========================================================================================================
#{String Utilities
#========================================================================================================
def encode(string):
    ''' Encode a unicode string
        @type     string:  C{unicode}
        @param    string:  Unicode string
        @rtype:   C{str}
        @return:  Encoded string
    '''
    if type(string) is unicode:return string.encode(encoding)
    else:return string

#========================================================================================================
#{Filesystem Utilities
#========================================================================================================
def archivelist(f):
    ''' List files in a tar (inc gzip or bz2 compressed) or zip archive.
        @type     f:  C{str}
        @param    f:  archive filepath
        @rtype:   C{list}
        @return:  archive filelisting
    '''
    lst=[]
    if tarfile.is_tarfile(f):
        #return tarfile.open(f,'r').getnames() #includes subfolders
        lst=[ti.name for ti in tarfile.open(f,'r').getmembers() if ti.isfile()]
        #return [os.sep.join(['/vsitar',normcase(f),l]) for l in lst]
        return [os.sep.join(['/vsitar',f,l]) for l in lst]

    elif zipfile.is_zipfile(f):
        #return zipfile.ZipFile(f,'r').namelist() #includes subfolders
        lst=[zi.filename for zi in zipfile.ZipFile(f,'r').infolist() if zi.file_size> 0]
        #return [os.sep.join(['/vsizip',normcase(f),l]) for l in lst]
        return [os.sep.join(['/vsizip',f,l]) for l in lst]
    return lst

def archivefileinfo(f,n):
    ''' List files in a tar (inc gzip or bz2 compressed) or zip archive.
        @type     f:  C{str}
        @param    f:  archive filepath
        @type     n:  C{str}
        @param    n:  archive member name
        @rtype:   C{dict}
        @return:  archive file member info
    '''
    archiveinfo={}

    if tarfile.is_tarfile(f):
        afi = tarfile.open(f,'r').getmember(n)
        archiveinfo['size']=afi.size
        archiveinfo['datemodified']=time.strftime(datetimeformat, time.localtime(afi.mtime))
        #archiveinfo['ownerid']=afi.uid  #Use the owner of the archive instead
        #archiveinfo['ownername']=afi.uname

    elif zipfile.is_zipfile(f):
        afi = zipfile.ZipFile(f,'r').getinfo(n)
        archiveinfo['size']=afi.file_size
        archiveinfo['datemodified']=time.strftime(datetimeformat, list(afi.date_time)+[0,0,0])

    return archiveinfo

def compressed_file_exists(path,testfile=True):
    ''' Check check whether /vsi...\path_to_archive\folder\file exists.
        Alternatively, only check if the archive exists on the file system.
        @type     path:      C{str}
        @param    path:      VSI filepath (/vsi...\path_to_archive\folder\file)
        @type     testfile:  C{bool}
        @param    testfile:  If True, check if file exists in archive. If False, only check if the archive exists on the file system.
        @rtype:   C{bool}
        @return:  Returns True or False
    '''
    p=os.path.split(path[8:])[0]
    while p:
        if os.path.exists(p) and tarfile.is_tarfile(p) or zipfile.is_zipfile(p):
            if testfile:
                if path in archivelist(p):return True
                else:return False
            else:return True

        p=os.path.split(p)[0]

    return False

def runcmd(cmd, format='s'):
    ''' Run a command
        @type     cmd:  C{str}
        @param    cmd:  Command (inc arguments) to run
        @rtype:   C{tuple}
        @return:  Returns (exit_code,stdout,stderr)
    '''
    import subprocess
    proc = subprocess.Popen(cmd, shell=True, stdin=subprocess.PIPE,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
    if format.lower() == 's': #string output
        stdout,stderr=proc.communicate()
    #elif format.lower() == 'f': #file object output #doesn't flush IO buffer, causes python to hang
    #    stdout,stderr=proc.stdout,proc.stderr
    elif format.lower() == 'l': #list output
        stdout,stderr=proc.stdout.readlines(),proc.stderr.readlines()
    #else:raise TypeError, "fomat argument must be in ['s','f','l'] (string, file, list)"
    else:raise TypeError, "fomat argument must be in ['s','l'] (string or list format)"
    exit_code=proc.wait()
    return exit_code,stdout,stderr

def which(name, returnfirst=True, flags=os.F_OK | os.X_OK, path=None):
    ''' Search PATH for executable files with the given name.

        On newer versions of MS-Windows, the PATHEXT environment variable will be
        set to the list of file extensions for files considered executable. This
        will normally include things like ".EXE". This fuction will also find files
        with the given name ending with any of these extensions.

        On MS-Windows the only flag that has any meaning is os.F_OK. Any other
        flags will be ignored.

        Derived mostly from U{http://code.google.com/p/waf/issues/detail?id=531} with
        additions from Brian Curtins patch - U{http://bugs.python.org/issue444582}

        @type name: C{str}
        @param name: The name for which to search.
        @type returnfirst: C{boolean}
        @param returnfirst: Return the first executable found.
        @type flags: C{int}
        @param flags: Arguments to U{os.access<http://docs.python.org/library/os.html#os.access>}.

        @rtype: C{str}/C{list}
        @return: Full path to the first matching file found or a list of the full paths to all files found,
                 in the order in which they were found.
    '''
    result = []
    exts = filter(None, os.environ.get('PATHEXT', '').split(os.pathsep))
    if not path:
        path = os.environ.get("PATH", os.defpath)
    for p in os.environ.get('PATH', '').split(os.pathsep):
        p = os.path.join(p, name)
        if os.access(p, flags):
            if returnfirst:return p
            else:result.append(p)
        for e in exts:
            pext = p + e
            if os.access(pext, flags):
                if returnfirst:return pext
                else:result.append(pext)
    return result

def exists(f,returnpath=False):
    ''' A case insensitive file existence checker

        @type f: C{str}
        @param f: The filepath to check.
        @type returnpath: C{boolean}
        @param returnpath: Return the case sensitive path.

        @rtype: C{boolean}/C{(str,boolean)}
        @return: True/False, optionally full path to the case sensitive path
    '''
    if iswin:#Windows is case insensitive anyways
        if returnpath:return os.path.exists(f),f
        else:return os.path.exists(f)
    import re
    path,name=os.path.split(os.path.abspath(f))
    files = os.listdir(path)
    for f in files:
        if re.search(f,name,re.I):
            if returnpath:return True,os.path.join(path,f)
            else:return True
    if returnpath:return False,None
    else:return False

def readbinary(data,offset, start, stop):
    ''' Read binary data
        @type    data:   C{str}
        @param   data:   data read from binary file
        @type    offset: C{int}
        @param   offset: Number of bytes to skip
        @type    start:  C{int}
        @param   start:  Byte to start reading from (from offset, not beginning of data)
        @type    stop:   C{int}
        @param   stop:   Byte to stop reading at (from offset, not beginning of data)
        @rtype:          C{str}
        @return:         String
    '''
    return ''.join(struct.unpack('s' * (stop-start+1), data[offset+start-1:offset+stop])).strip()

def readascii(data,offset,start,stop):
    ''' Read ASCII data
        @type    data:   C{str}
        @param   data:   data read from ASCII file
        @type    offset: C{int}
        @param   offset: Number of characters to skip
        @type    start:  C{int}
        @param   start:  Character to start reading from (from offset, not beginning of data)
        @type    stop:   C{int}
        @param   stop:   Character to stop reading at (from offset, not beginning of data)
        @rtype:          C{str}
        @return:         String
    '''
    return data[start+offset-1:stop+offset].strip()

def ByteOrder():
    ''' Determine byte order of host machine.

        @rtype:          C{str}
        @return:         String
    '''
    from struct import pack
    if pack('<h', 1) == pack('=h',1):
        return 'LSB'
    elif pack('>h', 1) == pack('=h',1):
        return 'MSB'
    else:
        raise Exception,'Unknown byte order'

def _WinFileOwner(filepath):
    import pywintypes
    import pythoncom
    import win32com.client
    import win32net
    import win32netcon

    OWNERID=(8,10) # seems to be 8 on XP, 10 on Win7
    try:
        d=os.path.split(filepath)
        oShell = win32com.client.Dispatch("Shell.Application")
        oFolder = oShell.NameSpace(d[0])
        for oid in OWNERID:
            ownerid=str(oFolder.GetDetailsOf(oFolder.parsename(d[1]), oid))
            if ownerid:break
        try:domain,ownerid=ownerid.split('\\')
        except:domain,ownerid=None,ownerid.split('\\')[-1]
    except: domain,ownerid=None,''
    #Too slow...
    ##oWMI = win32com.client.GetObject(r"winmgmts:\\.\root\cimv2")
    ##qry = "Select * from Win32_UserAccount where NAME = '%s'" % ownerid
    ##qry = oWMI.ExecQuery(qry)
    ##if qry.count > 0:
    ##for result in qry:
    ##    ownername=str(result.FullName)
    ##    break
    ##else: ownername='No user match'
    #Much quicker...
    try:
        dc=win32net.NetServerEnum(None,100,win32netcon.SV_TYPE_DOMAIN_CTRL)
        dcname=r'\\'+dc[0][0]['name']
    except:
        try:dcname=win32net.NetGetDCName()
        except:dcname=None
    try:
        if dcname:
            ownername=win32net.NetUserGetInfo(dcname,ownerid,2)['full_name']
        else:
           ownername=win32net.NetUserGetInfo(None,ownerid,2)['full_name']
    except: ownername='No user match'

    return ownerid,ownername

def _NixFileOwner(uid):
    import pwd
    pwuid=pwd.getpwuid(uid)
    ownerid = pwuid[0]
    ownername = pwuid[4]
    return ownerid,ownername

def FileInfo(filepath):
    ''' File information.

        @type    filepath: C{str}
        @param   filepath: Path to file
        @rtype:  C{dict}
        @return: Dictionary containing file: size, datemodified, datecreated, dateaccessed, ownerid & ownername
    '''
    fileinfo = {
        'size':0,
        'datemodified':'',
        'datecreated': '',
        'dateaccessed':'',
        'filepath':'',
        'guid':''
    }
    if not os.path.exists(filepath) and filepath[:4].lower()!= '/vsi':
        raise IOError('File not found')

    try:
        if filepath[:4].lower() == '/vsi':
            f=filepath.replace('/vsitar/','').replace('/vsitar\\','')
            f=f.replace('/vsizip/','').replace('/vsizip\\','')
            for ext in compressedfiles:
                if ext in f.lower():
                    f=f.split(ext)
                    archive=f[0]+ext
                    filename=ext.join(f[1:]).strip('\\/')
                    fileinfo.update(archivefileinfo(archive,filename))
                    break

            filestat = os.stat(archive)
            fileinfo['filename']=os.path.basename(filename)
            fileinfo['filepath']=filepath
            fileinfo['datecreated']=time.strftime(datetimeformat, time.localtime(filestat.st_ctime))
            fileinfo['dateaccessed']=time.strftime(datetimeformat, time.localtime(filestat.st_atime))
            fileinfo['guid']=uuid(filepath)
            filepath=archive
        else:
            #filepath=normcase(realpath(filepath))
            filepath=realpath(filepath)
            filestat = os.stat(filepath)

            fileinfo['filename']=os.path.basename(filepath)
            fileinfo['filepath']=filepath
            fileinfo['size']=filestat.st_size
            fileinfo['datemodified']=time.strftime(datetimeformat, time.localtime(filestat.st_mtime))
            fileinfo['datecreated']=time.strftime(datetimeformat, time.localtime(filestat.st_ctime))
            fileinfo['dateaccessed']=time.strftime(datetimeformat, time.localtime(filestat.st_atime))
            fileinfo['guid']=uuid(filepath)

        if not fileinfo.get('ownerid'):
            if iswin:
                ownerid,ownername=_WinFileOwner(filepath)
            else:
                ownerid,ownername=_NixFileOwner(filestat.st_uid)
            fileinfo['ownerid']=ownerid
            fileinfo['ownername']=ownername

    finally:return fileinfo


def uuid(filepath):
    ''' Generate a uuid reproducible based on filename

        @type    filepath: C{str}
        @param   filepath: Path to file
        @rtype:  C{str}
        @return: uuid
    '''
    #filepath=normcase(uncpath(realpath(filepath)))
    filepath=uncpath(realpath(filepath))
    return str(_uuid.uuid3(_uuid.NAMESPACE_DNS,filepath))

def uncpath(filepath):
    ''' Convert file path to UNC.

        @type    filepath: C{str}
        @param   filepath: Path to file
        @rtype:  C{str}
        @return: UNC filepath (if on Windows)
    '''
    #if sys.platform[0:3].lower()=='win':
    if iswin:
        import win32wnet
        if hasattr(filepath,'__iter__'): #Is iterable
            uncpath=[]
            for path in filepath:
                #try:    uncpath.append(normcase(win32wnet.WNetGetUniversalName(path)))
                #except: uncpath.append(normcase(path)) #Local path
                try:    uncpath.append(win32wnet.WNetGetUniversalName(path))
                except: uncpath.append(path) #Local path
        else:
            #try:    uncpath=win32wnet.WNetGetUniversalName(filepath)
            #except: uncpath=filepath #Local path
            try:    uncpath=win32wnet.WNetGetUniversalName(filepath)
            except: uncpath=filepath #Local path
    else:uncpath=filepath
    return uncpath

def normcase(filepath):
    ''' Normalize case of pathname. Makes all characters lowercase and all slashes into backslashes.

        @type    filepath: C{str/list}
        @param   filepath: Path to file/s
        @rtype:  C{str/list}
        @return: Path to file/s
    '''
    #if type(filepath) in [list,tuple]:
    if hasattr(filepath,'__iter__'): #Is iterable
        return [os.path.normcase(i) for i in filepath]
    else:
        return os.path.normcase(filepath)

def normpath(filepath):
    ''' Normalize path, eliminating double slashes, etc.

        @type    filepath: C{str/list}
        @param   filepath: Path to file/s
        @rtype:  C{str/list}
        @return: Path to file/s
    '''
    if hasattr(filepath,'__iter__'): #Is iterable
        return [os.path.normpath(i) for i in filepath]
    else:
        return os.path.normpath(filepath)

def realpath(filepath):
    ''' Return the absolute version of a path.

        @type    filepath: C{str/list}
        @param   filepath: Path to file/s
        @rtype:  C{str/list}
        @return: Path to file/s

        @note: os.path.realpath/os.path.abspath returns unexpected results on windows if filepath[-1]==':'
    '''
    if hasattr(filepath,'__iter__'): #Is iterable
        if iswin:
            realpath=[]
            for f in filepath:
                if f[-1]==':':f+='\\'
                realpath.append(os.path.realpath(f))
        else:return [os.path.realpath(f) for f in filepath]
    else:
        if iswin and filepath[-1]==':':filepath+='\\'
        return os.path.realpath(filepath)

def checkExt(filepath,ext):
    ''' Check a file has an allowed extension or apply a default extension if it has none.

        @type    filepath: C{str}
        @param   filepath: Path to file
        @type    ext: C{[str,...,str]}
        @param   ext: Allowed file extensions, ext[0] is the default
        @rtype:  C{str}
        @return: Path to file with updated extension
    '''
    vars=os.path.splitext(filepath)
    if vars[1] not in (ext):
        return vars[0]+ext[0]
    else:
        return filepath
def volname(path):
    ''' Get the volume label for a CD/DVD

        @type    path: C{str}
        @param   path: Disc path
        @rtype:  C{str}
        @return: Volume label
    '''
    volname=None
    try:
        #if sys.platform[0:3].lower()=='win':
        if iswin:
            import win32api
            drive=os.path.splitdrive(path)[0]
            if drive[-1]!='\\':drive+='\\'
            if drive: volinfo=win32api.GetVolumeInformation(drive)
            if volinfo[4] in ['CDFS','UDF']:volname=volinfo[0]
        else:
            #get the device from mount point
            exit_code,stdout,stderr=utilities.runcmd('df '+path)
            if exit_code == 0:
                device=stdout.split('\n')[1].split()[0]
                exit_code,stdout,stderr=runcmd('volname '+device)
                if exit_code == 0:volname=stdout.strip()
    finally:
        return volname

def writable(filepath):
    if not os.path.isdir(filepath):
        filepath=os.path.dirname(filepath)
    try:
        tmp=tempfile.TemporaryFile(dir=filepath) #Can we write a temp file there...?
        del tmp
        return True
    except:
        return False

class rglob:
    '''A recursive/regex enhanced glob
       adapted from os-path-walk-example-3.py - http://effbot.org/librarybook/os-path.htm
    '''
    def __init__(self, directory, pattern="*", regex=False, regex_flags=0, recurse=True, archive=False):
        ''' @type    directory: C{str}
            @param   directory: Path to xls file
            @type    pattern: C{type}
            @param   pattern: Regular expression/wildcard pattern to match files against
            @type    regex: C{boolean}
            @param   regex: Use regular expression matching (if False, use fnmatch)
                            See U{http://docs.python.org/library/re.html}
            @type    regex_flags: C{int}
            @param   regex_flags: Flags to pass to the regular expression compiler.
                                  See U{http://docs.python.org/library/re.html}
            @type    recurse: C{boolean}
            @param   recurse: Recurse into the directory?
            @type    archive: C{boolean}
            @param   archive: List files in compressed archives? Archive be supported by the zipfile and tarfile modules. Note: this slows things down considerably....
        '''
        self.stack = [directory]
        self.pattern = pattern
        self.regex = regex
        self.recurse = recurse
        self.archive = archive
        self.regex_flags = regex_flags
        self.files = []
        self.index = 0

    def __getitem__(self, index):
        while 1:
            try:
                file = self.files[self.index]
                self.index = self.index + 1
            except IndexError:
                # pop next directory from stack
                #self.directory = normcase(self.stack.pop())
                self.directory = self.stack.pop()
                try:
                    self.files = os.listdir(self.directory)
                    self.index = 0
                except:
                    if self.archive:
                        try:
                            self.files = archivelist(self.directory)
                            self.index = 0
                        except:pass
            else:
                # got a filename
                fullname = os.path.join(self.directory, file)
                try:islink=os.path.islink(fullname)
                except:islink=False
                try:isdir=os.path.isdir(fullname) and not islink
                except:isdir=False
                try:isarchive=(not islink and not isdir) and (tarfile.is_tarfile(fullname) or zipfile.is_zipfile(fullname))
                except:isarchive=False
                try:isfile=((not isdir and not isarchive and not islink) and os.path.isfile(fullname)) or (tarfile.is_tarfile(self.directory) or zipfile.is_zipfile(self.directory))
                except:isfile=False

                if isdir and self.recurse:
                    self.stack.append(fullname)
                elif isarchive and self.archive and os.path.exists(fullname):
                    self.stack.append(fullname)
                elif isfile:
                    if self.regex:
                        import re
                        if re.search(self.pattern,file,self.regex_flags):
                            return fullname
                    else:
                        import fnmatch
                        if fnmatch.fnmatch(file, self.pattern):
                            return fullname


#========================================================================================================
#{Process Utilities
#========================================================================================================
def isrunning(pid):
    if hasattr(os,'kill'):
        try:
            os.kill(pid, 0) #Sending a 0 signal does nothing.
            return True
        except:
            return False
    elif iswin:
        import win32process
        try:
            return pid in win32process.EnumProcesses()
        except:
            return False

#========================================================================================================
#{Exception Utilities
#========================================================================================================
def ExceptionInfo(maxTBlevel=0):
    '''Get info about the last exception'''
    cla, exc, trbk = sys.exc_info()
    excName = cla.__name__
    if maxTBlevel > 0:
        excArgs=[]
        excTb = FormatTraceback(trbk, maxTBlevel)
        #return '%s: %s\nTraceback: %s' % (excName, str(exc), excTb)
        return '%s: %s\n%s' % (excName, str(exc), excTb)
    else:
        return '%s: %s' % (excName, str(exc))

def FormatTraceback(trbk, maxTBlevel):
    '''Format traceback'''
    return 'Traceback (most recent call last): '+''.join(traceback.format_tb(trbk, maxTBlevel))

#========================================================================================================
#{Excel Utilities
#========================================================================================================
class ExcelWriter:
    ''' A simple spreadsheet writer'''

    def __init__(self,xlsx,fields=[],update=False, sort = True):
        ''' A simple spreadsheet writer.

            @type    xlsx: C{str}
            @param   xlsx: Path to xlsx file
            @type    fields: C{list}
            @param   fields: List of column/field headers
        '''

        if sort:fields.sort()
        self._file=xlsx
        self._fields=fields
        self._sheets=[]
        self._rows=1   #row index
        self._cols={}  #dict of col indices

        self._heading = openpyxl.styles.Style(font=openpyxl.styles.Font(bold=True))

        if update and os.path.exists(xlsx):
            self._wb=openpyxl.load_workbook(xlsx)
            self._sheets=self._wb.worksheets
            self._wb.encoding=encoding #
            self._ws=self._sheets[0]
            self._rows=self._ws.max_row-1

            #Check if all fields exist, add them if not
            ws=self._sheets[0]
            fields=[encode(c.value) for c in self._sheets[0].rows[0]]
            extrafields=[f for f in self._fields if f not in fields]
            col=len(fields)
            if extrafields:
                for ws in self._sheets:
                    #self._rows+=ws.max_row-1
                    row=ws.rows[0]
                    for i,field in enumerate(extrafields):
                        row[col+i].value=field
                fields+=extrafields
                self._wb.save(self._file)
            self._fields=fields

        else:
            if os.path.exists(xlsx):os.remove(xlsx)
            self._wb = openpyxl.Workbook(encoding=encoding)
            self._sheets = self._wb.worksheets
            self._ws = self._sheets[0]
            self._rows = 0
            self._addheader(self._ws)
            self._wb.save(self._file)

        #fs=set(self._fields) !!! set(list) reorders the list!!!
        fs=[]
        for f in self._fields:
            if f not in fs:fs.append(f)

        self._cols=dict(zip(fs,[self.__getcol__(self._fields,f) for f in fs]))

    def __getcol__(self,lst,val):
        i = -1
        cols=[]
        try:
            while 1:
                i = list(lst).index(val, i+1)
                cols.append(i)
        except ValueError:
            pass
        return cols

    def _addsheet(self):
        self._ws = self._wb.create_sheet()
        self._sheets=self._wb.worksheets
        self._addheader(self._ws)
        self._rows = 0

    def _addheader(self, ws):
        for i,field in enumerate(self._fields):
            ws.cell(row=1, column=i+1).value = field
            ws.cell(row=1, column=i+1).style = self._heading

    def _writevalue(self,row,col,value,ws=None):
        ''' Write a value to a cell

            @type    col: C{int}
            @param   col: column index, 0 based
            @type    row: C{int}
            @param   row: row index, 0 based
            @type    value: C{int/str}
            @param   value: value to write
        '''
        if not ws:ws=self._ws
        if isinstance(value,str):value=value.decode(encoding)
        if isinstance(value,basestring) and  len(value) > 32767:
            value=value[:32767]
            warnings.warn('The "%s" field is longer than 32767 characters and has been truncated.'%self._fields[field])
        ws.cell(row=row+1, column=col+1).value = value

    def WriteRecord(self,data):
        ''' Write a record

            @type    data: C{dict} #Known issue, doesn't handle list of lists (zipped lists)
            @param   data: Dict containing column headers (dict.keys()) and values (dict.values())
        '''
        dirty=False
        if self._rows > 1048575:
            self._addsheet()

        cols=copy.deepcopy(self._cols) #make a copy to alter
        if data!=dict(data):
            fields,values = zip(*data)
            for i,field in enumerate(fields):
                value=values[i]
                if field in self._fields and value not in ['',None,False]:#0 is valid
                    try:col=cols[field].pop(0)
                    except:continue
                    self._writevalue(self._rows+1, col,value)
                    dirty=True

        else:
            for field in data:
                if field in self._fields and data[field] not in ['',None,False]:#0 is valid
                    self._writevalue(self._rows+1, self._cols[field][0],data[field])
                    dirty=True

        if dirty:self._rows+=1
        self._wb.save(self._file)

    def UpdateRecord(self,data,row):
        ''' Update an existing record

            @type    data: C{dict} or C{list}
            @param   data: Dict containing column headers (dict.keys()) and values (dict.values()) or zipped list
            @type    row:  C{int}
            @param   row:  Row number of existing record
        '''
        dirty=False
        s=row/1048575
        r=row-s*1048575
        ws=self._wb.get_sheet(s)
        cols=copy.deepcopy(self._cols) #make a copy to alter
        if data!=dict(data):
            fields,values = zip(*data)
            for i,field in enumerate(fields):
                value=values[i]
                if field in self._fields and value not in ['',None,False]:#0 is valid
                    try:col=cols[field].pop(0)
                    except:continue
                    self._writevalue(r+1, col,values[i], ws)
                    dirty=True

        else:
            for field in data:
                if field in self._fields and data[field] not in ['',None,False]:#0 is valid
                    self._writevalue(r+1, self._cols[field][0],data[field], ws)
                    dirty=True
        if dirty:self._wb.save(self._file)

    def __del__(self):
        try:
            self._wb.save(self._file)
            del self._ws
            del self._wb
        except:pass


class ExcelReader:
    '''A simple spreadsheet reader'''
    def __init__(self,xlsx,returntype=dict):
        ''' A simple spreadsheet reader.

            @type    xlsx: C{str}
            @param   xlsx: Path to xlsx file
            @type    returntype: C{type}
            @param   returntype: dict or list
        '''
        self._wb=openpyxl.load_workbook(xlsx)
        self._returntype=returntype
        self._sheets=self._wb.worksheets
        self.records=0-len(self._sheets)
        self.headers=[encode(c.value) for c in self._sheets[0].rows[0]]
        for ws in self._sheets:
            self.records+=ws.max_row

    def __getitem__(self, index):
        i=index/1048575
        j=index-i*1048575
        ws=self._sheets[i]
        headers=[encode(c.value) for c in ws.rows[0]]
        cells=[encode(c.value) for c in ws.rows[j+1]]
        if self._returntype is dict:
            return dict(zip(headers,cells))
        else:
            return zip(headers,cells)

#}
